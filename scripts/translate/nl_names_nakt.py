#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Заполняет пустые поля NL в plants.csv на основе Excel-файла Naktuinbouw_Standaardlijst.xlsx.

Из Excel берутся:
- латинские названия из колонки "WETENSCHAPPELIJKE NAAM"
- нидерландские названия из колонок "NEDERLANDSE NAAM (voorkeur)" и "ALTERNATIEVE NAMEN"

Пример запуска:
    py -3 nl_names_nakt.py plants.csv Naktuinbouw_Standaardlijst.xlsx

Зависимости:
    pip install openpyxl
"""

import sys
import csv
import time
import shutil
from pathlib import Path
from collections import defaultdict

def norm_space(s: str) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split())

def norm_sci(s: str) -> str:
    return norm_space(s).lower()

def sniff_delimiter(path: Path, fallback=","):
    sample = path.read_text(encoding="utf-8-sig", errors="replace")[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",",";","\t","|"])
        return dialect.delimiter
    except Exception:
        return fallback

def backup_file(path: Path) -> Path:
    ts = time.strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
    shutil.copy2(str(path), str(backup))
    return backup

def normalize_header(h: str) -> str:
    # приводим к нижнему регистру и схлопываем пробелы/переводы строк
    h = (h or "").replace("\ufeff", "")
    h = " ".join(h.strip().split())
    return h.lower()

def split_alt_names(s: str):
    """
    Делит строку альтернативных названий по запятым, точкам с запятой, вертикальным чертам и слэшу.
    """
    if not s:
        return []
    raw = s.replace(";", ",").replace("|", ",").replace("/", ",")
    parts = [norm_space(p) for p in raw.split(",")]
    return [p for p in parts if p]

def load_nakt_map(xlsx_path: Path):
    """
    Читает первый лист Excel и строит маппинг:
        scientific_key -> set(dutch_names)
    Ожидаемые колонки (без учёта регистра/пробелов):
        'WETENSCHAPPELIJKE NAAM'
        'NEDERLANDSE NAAM (voorkeur)'
        'ALTERNATIEVE NAMEN'
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Ошибка: не установлен пакет openpyxl. Установите: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # Считываем заголовки
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(normalize_header(cell))

    # Индексы нужных колонок
    def col_idx(name):
        try:
            return headers.index(normalize_header(name))
        except ValueError:
            return -1

    idx_sci = col_idx("WETENSCHAPPELIJKE NAAM")
    idx_nl_pref = col_idx("NEDERLANDSE NAAM (voorkeur)")
    idx_nl_alt  = col_idx("ALTERNATIEVE NAMEN")

    if min(idx_sci, idx_nl_pref, idx_nl_alt) < 0:
        print("Ошибка: в Excel не найдены нужные колонки: "
              "'WETENSCHAPPELIJKE NAAM', 'NEDERLANDSE NAAM (voorkeur)', 'ALTERNATIEVE NAMEN'")
        sys.exit(1)

    names_map = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Безопасная выборка ячеек по индексам
        def at(idx):
            if idx < 0:
                return ""
            return "" if idx >= len(headers) else (row[idx] if idx < len(row) else "")
        sci = norm_sci(at(idx_sci))
        if not sci:
            continue

        pref = norm_space(at(idx_nl_pref))
        if pref:
            names_map[sci].add(pref)

        alt = norm_space(at(idx_nl_alt))
        for name in split_alt_names(alt):
            names_map[sci].add(name)

    return names_map

def main():
    if len(sys.argv) != 3:
        print("Usage: py -3 nl_names_nakt.py plants.csv Naktuinbouw_Standaardlijst.xlsx")
        sys.exit(1)

    plants_path = Path(sys.argv[1])
    xlsx_path   = Path(sys.argv[2])

    if not plants_path.exists():
        print(f"Не найден файл: {plants_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"Не найден файл: {xlsx_path}")
        sys.exit(1)

    names_map = load_nakt_map(xlsx_path)
    print(f"[nakt] Найдено латинских ключей: {len(names_map)}")

    # читаем plants.csv
    delim_plants = sniff_delimiter(plants_path, fallback=",")
    print(f"[plants] Использую разделитель: {repr(delim_plants)}")
    with open(plants_path, "r", encoding="utf-8-sig", newline="") as f:
        dr = csv.DictReader(f, delimiter=delim_plants)
        rows = list(dr)
        fieldnames = dr.fieldnames or []

    # карта нормализованное имя -> оригинальное
    colmap = { normalize_header(h): h for h in fieldnames }

    # проверим обязательные колонки
    required = ["sci", "nl"]
    missing = [r for r in required if r not in colmap]
    if missing:
        print("Ошибка: в plants.csv отсутствуют колонки: " + ", ".join(missing))
        sys.exit(1)

    # бэкап
    backup = backup_file(plants_path)
    print(f"Создан бэкап: {backup.name}")

    updated_nl = 0

    for row in rows:
        sci_raw = row.get(colmap["sci"], "")
        key = norm_sci(sci_raw)

        is_nl_empty = (row.get(colmap["nl"]) or "").strip() == ""
        if is_nl_empty and key in names_map and names_map[key]:
            row[colmap["nl"]] = " | ".join(sorted(names_map[key], key=str.lower))
            updated_nl += 1

    # записываем обратно с теми же заголовками и разделителем
    with open(plants_path, "w", encoding="utf-8-sig", newline="") as f:
        dw = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delim_plants)
        dw.writeheader()
        for row in rows:
            dw.writerow(row)

    print(f"Обновлено NL: {updated_nl}")
    print("Готово.")

if __name__ == "__main__":
    main()
